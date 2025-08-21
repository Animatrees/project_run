from typing import Type, Any

from django.core.files.uploadedfile import UploadedFile
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from rest_framework.serializers import Serializer


def import_xlsx_with_serializer(
        uploaded_file: UploadedFile,
        serializer_class: Type[Serializer],
        fields: list[str],
) -> list[list[Any]]:

    failed = []

    try:
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    except InvalidFileException as e:
        raise ValueError("Передан файл не в формате XLSX или повреждён") from e

    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_dict = dict(zip(fields, row))
            serializer = serializer_class(data=row_dict)
            if serializer.is_valid():
                serializer.save()
            else:
                failed.append(list(row))
    finally:
        wb.close()

    return failed
